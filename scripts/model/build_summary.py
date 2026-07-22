# -*- coding: utf-8 -*-
"""
Tính `08_Metrics_Summary` từ `fact_metrics_daily` + calendar + KPI target.

Chỗ dễ sai nhất và cũng là lý do file này tồn tại: **baseline**. Luật của kênh là so với
median 10 nội dung gần nhất CÙNG định dạng của chính kênh mình. Khi chưa đủ 10 mẫu,
script để TRỐNG và ghi "chưa đủ mẫu" — không thay bằng số ước lượng, không mượn benchmark
ngành. Một con số sai còn nguy hiểm hơn một ô trống.

    build_summary.py --calendar <calendar.csv> --metrics <fact_metrics_daily.csv>
                     --out <metrics_summary.csv> [--brief brief.yml] [--workbook <wb.xlsx>]
                     [--min-baseline-n 10]
"""
from __future__ import annotations

import argparse
import csv
import statistics
import sys
from collections import defaultdict
from datetime import date
from pathlib import Path

# workbook là nguồn sự thật; lib dùng chung để mọi script đọc giống nhau
sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "lib"))
from campaign_io import load_brief, load_calendar  # noqa: E402


try:
    import yaml
except ImportError:
    yaml = None

OUT_COLS = ["asset_id", "platform", "format", "pillar_primary", "first_date", "days_measured",
            "m24h_views", "m72h_views", "m7d_views", "m28d_views",
            "impressions", "reach", "views", "engaged_views", "engagement_total",
            "clicks_total", "cta_clicks", "subscribers_gained",
            "leads", "lead_per_1000_reach",
            "engagement_rate", "ctr", "cta_rate", "subs_per_1000_engaged", "retention_30s",
            "baseline_n", "vs_median_primary", "vs_median_secondary",
            "kpi_hit_count", "guardrail_ok", "is_winner", "decision_28d", "decision_note"]

SUMMABLE = ["impressions", "reach", "views", "engaged_views", "engagement_total",
            "clicks_total", "cta_clicks", "subscribers_gained", "likes", "comments", "shares"]
AVERAGEABLE = ["retention_30s", "avg_percentage_viewed", "ctr"]


def num(v):
    if v in (None, "", "-"):
        return None
    try:
        f = float(str(v).replace(",", ""))
        return f
    except ValueError:
        return None


def read_csv(p: str) -> list[dict]:
    with open(p, encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def ratio(a, b):
    """a/b, trả None khi thiếu dữ liệu hoặc mẫu = 0 — KHÔNG trả 0."""
    if a is None or b in (None, 0):
        return None
    return round(a / b, 6)


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--calendar", required=True)
    ap.add_argument("--metrics", required=True)
    ap.add_argument("--out", required=True)
    ap.add_argument("--brief")
    ap.add_argument("--workbook")
    ap.add_argument("--leads-bridge", help="bảng cầu post_id ↔ lead_id để quy lead về asset")
    ap.add_argument("--min-baseline-n", type=int, default=10)
    args = ap.parse_args()

    cal = {r["asset_id"]: r for r in load_calendar(args.calendar) if r.get("asset_id")}
    facts = read_csv(args.metrics)

    primary, secondary, guardrail, min_hits = "views", "engagement_rate", "retention_30s", 2
    if args.brief and Path(args.brief).exists():
        b = load_brief(args.brief)
        primary = b.get("primary_kpi") or primary
        secondary = b.get("secondary_kpi") or secondary
        guardrail = b.get("guardrail_kpi") or guardrail
        min_hits = ((b.get("winner_rule") or {}).get("min_kpi_count")) or min_hits

    # Lead quy về asset qua bảng cầu CTA. Không có bảng này thì cột leads để TRỐNG,
    # không phải 0 — "chưa nối được" khác hẳn "có 0 lead".
    lead_count: dict[str, int] | None = None
    if args.leads_bridge and Path(args.leads_bridge).exists():
        lead_count = defaultdict(int)
        for b in read_csv(args.leads_bridge):
            pid = (b.get("post_id") or b.get("asset_id") or "").strip()
            lid = (b.get("lead_id") or "").strip()
            if pid and lid:
                lead_count[pid] += 1
        print(f"   lead bridge: {sum(lead_count.values())} lead quy về {len(lead_count)} asset")

    # gom theo (asset, platform)
    grp: dict[tuple, list[dict]] = defaultdict(list)
    for f in facts:
        grp[(f["asset_id"], f["platform"])].append(f)

    rows = []
    for (aid, plat), recs in sorted(grp.items()):
        recs.sort(key=lambda r: r["date"])
        first = date.fromisoformat(recs[0]["date"])
        agg = {c: None for c in OUT_COLS}
        c = cal.get(aid, {})
        agg.update({"asset_id": aid, "platform": plat,
                    "format": c.get("format", ""), "pillar_primary": c.get("pillar_primary", ""),
                    "first_date": recs[0]["date"], "days_measured": len(recs)})

        for m in SUMMABLE:
            vals = [num(r.get(m)) for r in recs]
            vals = [v for v in vals if v is not None]
            if vals:
                agg[m] = round(sum(vals), 4)
        for m in AVERAGEABLE:
            vals = [num(r.get(m)) for r in recs]
            vals = [v for v in vals if v is not None]
            if vals:
                agg[m] = round(statistics.mean(vals), 6)

        # mốc thời gian — cộng dồn tới ngày thứ N kể từ ngày đăng
        for label, ndays in (("m24h_views", 1), ("m72h_views", 3),
                             ("m7d_views", 7), ("m28d_views", 28)):
            vals = [num(r.get("views")) for r in recs
                    if (date.fromisoformat(r["date"]) - first).days < ndays]
            vals = [v for v in vals if v is not None]
            agg[label] = round(sum(vals), 2) if vals else None

        agg["engagement_rate"] = ratio(agg.get("engagement_total"), agg.get("reach"))
        agg["ctr"] = ratio(agg.get("clicks_total"), agg.get("impressions"))
        agg["cta_rate"] = ratio(agg.get("cta_clicks"), agg.get("reach"))
        sp = ratio(agg.get("subscribers_gained"), agg.get("engaged_views"))
        agg["subs_per_1000_engaged"] = round(sp * 1000, 4) if sp is not None else None
        # Lead gắn theo ASSET, còn bảng này grain (asset × kênh). Nếu điền cho mọi dòng
        # thì một lead bị đếm hai lần khi asset chạy trên hai kênh. Chỉ quy về kênh chính;
        # dòng kênh phụ để TRỐNG (chưa quy được), không phải 0.
        if lead_count is not None and plat == (c.get("platform") or plat):
            agg["leads"] = lead_count.get(aid, 0)
            lp = ratio(agg["leads"], agg.get("reach"))
            agg["lead_per_1000_reach"] = round(lp * 1000, 4) if lp is not None else None
        rows.append(agg)

    # ── baseline: median theo ĐỊNH DẠNG, chỉ dùng khi đủ mẫu ──────────────────
    by_fmt: dict[str, list[dict]] = defaultdict(list)
    for r in rows:
        by_fmt[r["format"] or "?"].append(r)

    def median_of(fmt: str, metric: str, exclude: str) -> tuple[float | None, int]:
        vals = [r[metric] for r in by_fmt[fmt]
                if r["asset_id"] != exclude and r.get(metric) is not None]
        return (statistics.median(vals) if vals else None), len(vals)

    # KPI trong brief phải là TÊN CỘT canonical, không phải nhãn tiếng Việt.
    # Nếu không map được thì không tính được winner — phải kêu to, không được lặng lẽ bỏ qua.
    computable = [c for c in OUT_COLS if c not in
                  ("asset_id", "platform", "format", "pillar_primary", "first_date",
                   "days_measured", "baseline_n", "vs_median_primary", "vs_median_secondary",
                   "kpi_hit_count", "guardrail_ok", "is_winner", "decision_28d", "decision_note")]
    bad_kpi = [(lbl, k) for lbl, k in (("primary_kpi", primary), ("secondary_kpi", secondary),
                                       ("guardrail_kpi", guardrail))
               if k not in computable]
    if bad_kpi:
        print("   ✖ KPI trong brief KHÔNG phải tên cột đo được — không kết luận winner được:",
              file=sys.stderr)
        for lbl, k in bad_kpi:
            print(f"       {lbl} = '{k}'", file=sys.stderr)
        print(f"     Phải dùng một trong: {', '.join(computable)}", file=sys.stderr)
        print("     (Nhãn hiển thị tiếng Việt để ở primary_kpi_label, không để ở primary_kpi.)",
              file=sys.stderr)

    thin: set[str] = set()
    for r in rows:
        fmt = r["format"] or "?"
        hits, n_ref = 0, 0
        for kpi, col in ((primary, "vs_median_primary"), (secondary, "vs_median_secondary")):
            key = kpi if kpi in r else None
            if key is None:
                continue
            med, n = median_of(fmt, key, r["asset_id"])
            n_ref = max(n_ref, n)
            if med is None or n < args.min_baseline_n or not med:
                r[col] = None                      # ← chưa đủ mẫu: ĐỂ TRỐNG
                thin.add(fmt)
                continue
            r[col] = round(r[key] / med - 1, 4) if r.get(key) is not None else None
            if r[col] is not None and r[col] > 0:
                hits += 1
        r["baseline_n"] = n_ref
        r["kpi_hit_count"] = hits
        gmed, gn = median_of(fmt, guardrail, r["asset_id"]) if guardrail in r else (None, 0)
        if gmed is None or gn < args.min_baseline_n or r.get(guardrail) is None:
            r["guardrail_ok"] = ""                 # chưa kết luận được
            r["is_winner"] = ""
        else:
            r["guardrail_ok"] = "TRUE" if r[guardrail] >= gmed else "FALSE"
            r["is_winner"] = "TRUE" if (hits >= min_hits and r["guardrail_ok"] == "TRUE") else "FALSE"

    out = Path(args.out)
    out.parent.mkdir(parents=True, exist_ok=True)
    with open(out, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=OUT_COLS, extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)

    print(f"── tổng hợp · {len(rows)} dòng (asset × kênh)")
    print(f"   KPI: {primary} (chính) · {secondary} (phụ) · {guardrail} (guardrail) "
          f"· cần ≥{min_hits} KPI vượt median")
    fmt_n = {f: len(v) for f, v in sorted(by_fmt.items())}
    print("   mẫu theo định dạng: " + " · ".join(f"{k}={v}" for k, v in fmt_n.items()))
    decided = [r for r in rows if r["is_winner"] in ("TRUE", "FALSE")]
    if decided:
        print(f"   kết luận winner được cho {len(decided)}/{len(rows)} dòng "
              f"({sum(1 for r in decided if r['is_winner'] == 'TRUE')} winner)")
    if thin:
        print(f"   ⚠ CHƯA ĐỦ MẪU (cần ≥{args.min_baseline_n} nội dung cùng định dạng) → "
              f"để trống vs_median cho: {', '.join(sorted(thin))}")
        print(f"     Đây là hành vi đúng: thà trống còn hơn một con số không kiểm được.")
    print(f"   ✅ {out}")

    if args.workbook:
        try:
            from openpyxl import load_workbook
        except ImportError:
            print("   ⚠ thiếu openpyxl — bỏ qua nạp workbook")
            return 0
        wb = load_workbook(args.workbook)
        ws = wb["08_Metrics_Summary"]
        hdr = {}
        for i, cell in enumerate(ws[1], start=1):
            n = str(cell.value or "")
            for m in ("🤖", "👤", "⚙️", "⚙"):
                n = n.replace(m, "")
            hdr[n.strip()] = i
        ws.delete_rows(2, ws.max_row)
        for j, r in enumerate(rows, start=2):
            for name, i in hdr.items():
                if name in r and r[name] not in (None, ""):
                    ws.cell(j, i, r[name])
        wb.save(args.workbook)
        print(f"   ✅ nạp {len(rows)} dòng vào 08_Metrics_Summary")
    return 0


if __name__ == "__main__":
    sys.exit(main())
