# -*- coding: utf-8 -*-
"""
Nạp số liệu từ file export tay về `fact_metrics_daily`.

Đây là đường KHÔNG cần quyền API: bạn tải CSV từ Meta Business Suite và YouTube Studio,
script map về schema canonical. Cũng chính là bài lab S04 của khoá MKA-101.

Ánh xạ tên cột đọc từ `schema/crosswalk.yml` — không hardcode ở đây, nên khi Facebook
đổi tên cột thì sửa crosswalk là xong.

    from_export.py --calendar <calendar.csv> --out <fact_metrics_daily.csv>
                   [--facebook-daily "Facebook Post Daily.csv"]
                   [--youtube youtube_export.csv]
                   [--workbook <wb.xlsx>]      # nạp luôn vào sheet 07_Metrics_Daily

Exit: 0 ok · 1 có dòng không map được asset · 2 không chạy được.
"""
from __future__ import annotations

import argparse
import csv
import sys
from datetime import datetime
from pathlib import Path

try:
    import yaml
except ImportError:
    print("Thiếu pyyaml: pip install pyyaml", file=sys.stderr)
    sys.exit(2)

SCHEMA_DIR = Path(__file__).resolve().parents[2] / "schema"

OUT_COLS = ["asset_id", "platform", "date", "source", "impressions", "ctr", "reach",
            "reach_organic", "reach_paid", "views", "engaged_views",
            "avg_view_duration_sec", "avg_percentage_viewed", "watch_time_min",
            "retention_30s", "likes", "comments", "shares", "engagement_total",
            "poll_votes", "subscribers_gained", "related_video_clicks",
            "returning_viewers", "clicks_total", "cta_clicks", "leads", "cost"]

# Cột Facebook export (tiếng Việt) → canonical. Nguồn: schema/crosswalk.yml
FB_METRIC = {
    "Lượt xem": "impressions",
    "Số người tiếp cận": "reach",
    "Số người tiếp cận từ Bài viết tự nhiên": "reach_organic",
    "Số người tiếp cận từ Bài viết đã quảng cáo": "reach_paid",
    "Cảm xúc": "likes",
    "Bình luận": "comments",
    "Lượt chia sẻ": "shares",
    "Tổng lượt click": "clicks_total",
    "Lượt click vào liên kết": "cta_clicks",
    "Lượt xem video trong tối thiểu 3 giây": "views",
}
YT_METRIC = {
    "impressions": "impressions", "ctr": "ctr", "views": "views",
    "engaged_views": "engaged_views", "avg_view_duration_sec": "avg_view_duration_sec",
    "avg_percentage_viewed": "avg_percentage_viewed", "watch_time_min": "watch_time_min",
    "likes": "likes", "comments": "comments", "shares": "shares",
    "subscribers_gained": "subscribers_gained",
}


def num(v):
    """Ép về số; rỗng/không parse được → None (KHÔNG phải 0 — rỗng nghĩa là chưa biết)."""
    if v in (None, "", "-"):
        return None
    s = str(v).strip().replace(",", "").replace("%", "")
    try:
        f = float(s)
        return int(f) if f.is_integer() else f
    except ValueError:
        return None


def parse_date(v: str) -> str | None:
    """Facebook export dùng MM/DD/YYYY; YouTube Studio dùng ISO. Nhận cả hai."""
    if not v:
        return None
    s = str(v).strip().split(" ")[0]
    for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def read_csv(p: Path) -> list[dict]:
    with open(p, encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def build_index(cal_rows: list[dict]) -> dict[str, tuple[str, str]]:
    """native_id → (asset_id, platform). Nhận cả cột canonical lẫn cột tách theo kênh."""
    idx: dict[str, tuple[str, str]] = {}
    for r in cal_rows:
        aid = (r.get("asset_id") or "").strip()
        if not aid:
            continue
        for col, plat in (("platform_native_id", (r.get("platform") or "").strip()),
                          ("platform_native_id_fb", "facebook"),
                          ("platform_native_id_yt", "youtube")):
            v = (r.get(col) or "").strip()
            if v:
                idx[v] = (aid, plat or "facebook")
    return idx


def collect_facebook(rows: list[dict], idx: dict, unmapped: set) -> list[dict]:
    out = []
    for r in rows:
        nid = str(r.get("ID bài viết") or "").strip()
        hit = idx.get(nid)
        if not hit:
            if nid:
                unmapped.add(nid)
            continue
        d = parse_date(r.get("Ngày") or r.get("Thời gian đăng"))
        if not d:
            continue
        rec = {c: None for c in OUT_COLS}
        rec.update({"asset_id": hit[0], "platform": "facebook", "date": d,
                    "source": "manual_export"})
        for src, dst in FB_METRIC.items():
            if src in r:
                rec[dst] = num(r[src])
        parts = [rec["likes"], rec["comments"], rec["shares"]]
        if any(p is not None for p in parts):
            rec["engagement_total"] = sum(p or 0 for p in parts)
        out.append(rec)
    return out


def collect_youtube(rows: list[dict], idx: dict, unmapped: set) -> list[dict]:
    out = []
    for r in rows:
        nid = str(r.get("video_id") or r.get("Video") or "").strip()
        hit = idx.get(nid)
        if not hit:
            if nid:
                unmapped.add(nid)
            continue
        d = parse_date(r.get("date") or r.get("publish_date") or r.get("Ngày"))
        if not d:
            continue
        rec = {c: None for c in OUT_COLS}
        rec.update({"asset_id": hit[0], "platform": "youtube", "date": d,
                    "source": "manual_export"})
        for src, dst in YT_METRIC.items():
            if src in r:
                rec[dst] = num(r[src])
        parts = [rec["likes"], rec["comments"], rec["shares"]]
        if any(p is not None for p in parts):
            rec["engagement_total"] = sum(p or 0 for p in parts)
        out.append(rec)
    return out


def dedupe(rows: list[dict]) -> list[dict]:
    """Khoá (asset_id, platform, date) — bản nạp sau đè bản trước."""
    seen: dict[tuple, dict] = {}
    for r in rows:
        seen[(r["asset_id"], r["platform"], r["date"])] = r
    return sorted(seen.values(), key=lambda r: (r["asset_id"], r["platform"], r["date"]))


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--calendar", required=True)
    ap.add_argument("--out", required=True)
    ap.add_argument("--facebook-daily")
    ap.add_argument("--youtube")
    ap.add_argument("--workbook")
    args = ap.parse_args()

    if not (SCHEMA_DIR / "crosswalk.yml").exists():
        print("Không thấy schema/crosswalk.yml", file=sys.stderr)
        return 2

    cal = read_csv(Path(args.calendar))
    idx = build_index(cal)
    print(f"── nạp số liệu · {len(cal)} asset · {len(idx)} khoá nền tảng")
    if not idx:
        print("   ✖ Calendar chưa có platform_native_id nào — chưa đăng thì chưa có gì để đo.",
              file=sys.stderr)
        return 1

    rows: list[dict] = []
    unmapped: set[str] = set()
    if args.facebook_daily:
        fb = read_csv(Path(args.facebook_daily))
        got = collect_facebook(fb, idx, unmapped)
        print(f"   facebook: {len(fb)} dòng export → {len(got)} dòng khớp asset")
        rows += got
    if args.youtube:
        yt = read_csv(Path(args.youtube))
        got = collect_youtube(yt, idx, unmapped)
        print(f"   youtube : {len(yt)} dòng export → {len(got)} dòng khớp asset")
        rows += got

    if not rows:
        print("   ✖ Không nạp được dòng nào.", file=sys.stderr)
        return 1

    rows = dedupe(rows)
    out = Path(args.out)
    out.parent.mkdir(parents=True, exist_ok=True)
    with open(out, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=OUT_COLS)
        w.writeheader()
        w.writerows(rows)

    days = sorted({r["date"] for r in rows})
    assets = {r["asset_id"] for r in rows}
    print(f"   ✅ {out}")
    print(f"      {len(rows)} dòng · {len(assets)} asset · {days[0]} → {days[-1]}")
    if unmapped:
        print(f"   ⚠ {len(unmapped)} native_id trong export KHÔNG có trong calendar "
              f"(vd {', '.join(list(unmapped)[:3])}) — bài đăng ngoài chiến dịch, đã bỏ qua.")

    if args.workbook:
        try:
            from openpyxl import load_workbook
        except ImportError:
            print("   ⚠ thiếu openpyxl — bỏ qua bước nạp vào workbook")
            return 0
        wb = load_workbook(args.workbook)
        ws = wb["07_Metrics_Daily"]
        hdr = {}
        for i, c in enumerate(ws[1], start=1):
            n = str(c.value or "")
            for m in ("🤖", "👤", "⚙️", "⚙"):
                n = n.replace(m, "")
            hdr[n.strip()] = i
        ws.delete_rows(2, ws.max_row)
        for j, r in enumerate(rows, start=2):
            for name, i in hdr.items():
                if name in r and r[name] is not None:
                    ws.cell(j, i, r[name])
        wb.save(args.workbook)
        print(f"   ✅ nạp {len(rows)} dòng vào 07_Metrics_Daily của {Path(args.workbook).name}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
