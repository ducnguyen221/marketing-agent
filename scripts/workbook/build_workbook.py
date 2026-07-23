# -*- coding: utf-8 -*-
"""
Dựng file Excel quản lý chiến dịch — 5 sheet chuẩn KPIM (Campaign / Post / Result /
Engagement / Assets) từ schema/workbook_spec.yml.

Hai chế độ:
  init-template   dựng workbook rỗng làm mẫu (Campaign form + header các bảng)
  new-campaign    copy mẫu + đổ Sheet Campaign từ campaign_meta.json

    build_workbook.py init-template --out <tracking_template.xlsx>
    build_workbook.py new-campaign  --out <NN_Ten.xlsx> --meta <campaign_meta.json>
"""
from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

try:
    import yaml
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError as e:
    print(f"Thiếu thư viện: {e}. Cần: pip install pyyaml openpyxl", file=sys.stderr)
    sys.exit(2)

SPEC = Path(__file__).resolve().parents[2] / "schema" / "workbook_spec.yml"
HDR_FILL = PatternFill("solid", fgColor="1F4E79")
HDR_FONT = Font(bold=True, size=10, color="FFFFFF")
FORM_FILL = PatternFill("solid", fgColor="E2EFDA")
GATE_FILL = PatternFill("solid", fgColor="FFF2CC")   # 3 cột approve_* = của người
GATE_COLS = {"approve_topic", "approve_content", "approve_final"}


def load_spec() -> dict:
    return yaml.safe_load(SPEC.read_text(encoding="utf-8"))


def build(spec: dict) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    for sh in spec["sheets"]:
        ws = wb.create_sheet(sh["name"])
        if sh.get("kind") == "form":
            for i, h in enumerate(sh["columns"], start=1):
                c = ws.cell(1, i, h)
                c.font = HDR_FONT
                c.fill = HDR_FILL
            for r, (field, hint) in enumerate(sh["fields"], start=2):
                ws.cell(r, 1, field).font = Font(bold=True, size=10)
                ws.cell(r, 1).fill = FORM_FILL
                ws.cell(r, 3, hint).font = Font(italic=True, size=9, color="808080")
            for col, w in (("A", 22), ("B", 46), ("C", 50)):
                ws.column_dimensions[col].width = w
            for r in range(2, len(sh["fields"]) + 2):
                ws.cell(r, 2).alignment = Alignment(wrap_text=True, vertical="top")
        else:
            for i, col in enumerate(sh["columns"], start=1):
                c = ws.cell(1, i, col)
                c.font = HDR_FONT
                c.fill = GATE_FILL if col in GATE_COLS else HDR_FILL
                if col in GATE_COLS:
                    c.font = Font(bold=True, size=10, color="7F6000")
                ws.column_dimensions[get_column_letter(i)].width = \
                    max(len(col) + 2, 10 if col not in ("detail_prompt", "notes", "angle") else 30)
            ws.freeze_panes = sh.get("_freeze", "A2")
            ws.auto_filter.ref = f"A1:{get_column_letter(len(sh['columns']))}1"
        ws.row_dimensions[1].height = 22
    return wb


def cmd_init(args) -> int:
    wb = build(load_spec())
    out = Path(args.out)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"✅ template: {out}")
    print(f"   {len(wb.sheetnames)} sheet: {', '.join(wb.sheetnames)}")
    return 0


def cmd_new(args) -> int:
    spec = load_spec()
    wb = build(spec)
    meta = {}
    if args.meta and Path(args.meta).exists():
        meta = json.loads(Path(args.meta).read_text(encoding="utf-8"))
    ws = wb["Campaign"]
    field_row = {ws.cell(r, 1).value: r for r in range(2, ws.max_row + 1)}
    filled = 0
    for k, v in meta.items():
        if k in field_row:
            ws.cell(field_row[k], 2, v)
            filled += 1
    out = Path(args.out)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"✅ campaign workbook: {out}")
    print(f"   {len(wb.sheetnames)} sheet · đổ {filled} field Sheet Campaign từ meta")
    return 0


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    sub = ap.add_subparsers(dest="cmd", required=True)
    p = sub.add_parser("init-template"); p.add_argument("--out", required=True); p.set_defaults(fn=cmd_init)
    p = sub.add_parser("new-campaign"); p.add_argument("--out", required=True)
    p.add_argument("--meta"); p.set_defaults(fn=cmd_new)
    args = ap.parse_args()
    return args.fn(args)


if __name__ == "__main__":
    sys.exit(main())
