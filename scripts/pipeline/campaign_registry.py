# -*- coding: utf-8 -*-
"""
Sổ chiến dịch — quét mọi folder campaign của một instance, sinh:
  CAMPAIGNS.md    người đọc: đang chạy gì, tới đâu
  campaigns.xlsx  máy đọc: lịch sử để lọc/sắp xếp

Quét thẳng Sheet Campaign + Sheet Post của từng workbook nên không lệch thực tế.
Mô hình 5 sheet KPIM.

    campaign_registry.py --instance <content_root>
"""
from __future__ import annotations

import argparse
import sys
from collections import Counter
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "lib"))
from campaign_io import find_workbook, read_campaign, read_sheet  # noqa: E402

TRUTHY = {"x", "true", "1", "✓", "✔"}
COLS = ["campaign_code", "campaign_id", "name", "pillar", "status", "owner",
        "schedule_start", "schedule_end", "num_posts_planned",
        "posts_total", "posts_published", "gate_topic", "gate_content", "gate_final",
        "kpi_targets", "channels", "folder"]


def is_true(v):
    return str(v or "").strip().lower() in TRUTHY


def scan(inst: Path) -> list[dict]:
    cdir = inst / "02_campaigns"
    if not cdir.exists():
        return []
    rows = []
    for d in sorted(p for p in cdir.iterdir() if p.is_dir()):
        wb = find_workbook(d)
        if not wb:
            continue
        camp = read_campaign(wb)
        post = read_sheet(wb, "Post")
        st = Counter(str(r.get("status") or "?") for r in post)
        rows.append({
            "campaign_code": camp.get("campaign_code") or d.name,
            "campaign_id": camp.get("campaign_id", ""),
            "name": camp.get("name", ""),
            "pillar": camp.get("pillar", ""),
            "status": camp.get("status", ""),
            "owner": camp.get("owner", ""),
            "schedule_start": camp.get("schedule_start", ""),
            "schedule_end": camp.get("schedule_end", ""),
            "num_posts_planned": camp.get("num_posts_planned", ""),
            "posts_total": len(post),
            "posts_published": st.get("published", 0),
            "gate_topic": sum(1 for r in post if is_true(r.get("approve_topic"))),
            "gate_content": sum(1 for r in post if is_true(r.get("approve_content"))),
            "gate_final": sum(1 for r in post if is_true(r.get("approve_final"))),
            "kpi_targets": camp.get("kpi_targets", ""),
            "channels": camp.get("channels", ""),
            "folder": d.name,
        })
    return rows


def write_md(rows, out: Path, name: str):
    running = [r for r in rows if str(r["status"]).lower() in ("active", "running")]
    L = [f"# Sổ chiến dịch — {name}", "",
         "> Sinh tự động bởi `campaign_registry.py`. Đừng sửa tay — sửa Sheet Campaign của",
         "> workbook rồi chạy lại.", "",
         f"**{len(rows)} chiến dịch** · {sum(r['posts_total'] for r in rows)} bài · "
         f"{sum(r['posts_published'] for r in rows)} đã đăng", "",
         "## Đang chạy", ""]
    if running:
        L += ["| Chiến dịch | Pillar | Bài | Đã đăng | Cổng (T/C/F) |",
              "|---|---|---:|---:|---|"]
        for r in running:
            L.append(f"| **{r['campaign_code']}** — {r['name']} | {r['pillar']} "
                     f"| {r['posts_total']} | {r['posts_published']} "
                     f"| {r['gate_topic']}/{r['gate_content']}/{r['gate_final']} |")
    else:
        L.append("*Chưa có chiến dịch nào ở trạng thái `active`.*")
    L += ["", "## Toàn bộ", "",
          "| Chiến dịch | Tên | Pillar | Thời gian | Status | Bài | Đã đăng |",
          "|---|---|---|---|---|---:|---:|"]
    for r in sorted(rows, key=lambda x: str(x["schedule_start"]), reverse=True):
        L.append(f"| `{r['campaign_code']}` | {r['name']} | {r['pillar']} "
                 f"| {r['schedule_start']} → {r['schedule_end']} | {r['status']} "
                 f"| {r['posts_total']} | {r['posts_published']} |")
    L += ["", "> Cổng (T/C/F) = số bài đã tick approve_topic / approve_content / approve_final.", ""]
    out.write_text("\n".join(L), encoding="utf-8")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--instance", required=True)
    args = ap.parse_args()
    inst = Path(args.instance)
    rows = scan(inst)
    if not rows:
        print(f"Không thấy chiến dịch nào trong {inst / '02_campaigns'}", file=sys.stderr)
        return 1
    print(f"── sổ chiến dịch · {inst.name} · {len(rows)} chiến dịch")
    for r in rows:
        print(f"   {r['campaign_code']:<20} {str(r['status']):<10} {r['posts_total']:>3} bài "
              f"· cổng {r['gate_topic']}/{r['gate_content']}/{r['gate_final']}")
    write_md(rows, inst / "CAMPAIGNS.md", inst.name)
    print(f"   ✅ {inst / 'CAMPAIGNS.md'}")
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        wb = Workbook(); ws = wb.active; ws.title = "Campaigns"
        for i, c in enumerate(COLS, 1):
            cell = ws.cell(1, i, c); cell.font = Font(bold=True, size=10, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E79")
            ws.column_dimensions[cell.column_letter].width = max(len(c) + 2, 14)
        for j, r in enumerate(rows, 2):
            for i, c in enumerate(COLS, 1):
                ws.cell(j, i, r.get(c, ""))
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{ws.cell(1, len(COLS)).column_letter}{len(rows)+1}"
        wb.save(inst / "campaigns.xlsx")
        print(f"   ✅ {inst / 'campaigns.xlsx'}")
    except ImportError:
        pass
    return 0


if __name__ == "__main__":
    sys.exit(main())
