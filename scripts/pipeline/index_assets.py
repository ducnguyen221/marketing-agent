# -*- coding: utf-8 -*-
"""
Quét thư mục chiến dịch → điền `11_Assets` và sinh `12_Links` (UTM).

Hai việc hay bị bỏ quên nhất và trả giá muộn nhất:
  · file nằm đâu, ai làm, có quyền dùng không   → 11_Assets
  · click quy về asset nào                      → 12_Links

Quy ước thư mục (đặt trong thư mục chiến dịch):
    assets/images/<asset_id>-*.png|jpg|svg
    assets/video/<asset_id>-*.mp4
    assets/audio/<asset_id>-*.mp3
    drafts/<asset_id>-*.md

`asset_id` lấy từ phần đầu tên file trước dấu `-` cuối cùng khớp calendar.

    index_assets.py --campaign-dir <dir> --workbook <wb.xlsx>
                    [--utm-base https://example.com/lp] [--utm-source ...] [--dry-run]
"""
from __future__ import annotations

import argparse
import csv
import hashlib
import sys
from datetime import datetime
from pathlib import Path

# workbook là nguồn sự thật; lib dùng chung để mọi script đọc giống nhau
sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "lib"))
from campaign_io import load_brief, load_calendar  # noqa: E402


try:
    from openpyxl import load_workbook
except ImportError:
    print("Thiếu openpyxl: pip install openpyxl", file=sys.stderr)
    sys.exit(2)

KIND = {".png": "image", ".jpg": "image", ".jpeg": "image", ".svg": "image", ".webp": "image",
        ".mp4": "video", ".mov": "video", ".mp3": "audio", ".wav": "audio",
        ".md": "markdown", ".docx": "doc", ".pdf": "doc"}
ROLE_HINT = {"thumb": "thumbnail", "hero": "hero_media", "broll": "b_roll",
             "info": "infographic", "script": "script", "caption": "caption",
             "sub": "subtitle"}
UTM_MEDIUM = {"youtube": "video", "facebook": "social"}


def clean(name: str) -> str:
    for m in ("🤖", "👤", "⚙️", "⚙"):
        name = name.replace(m, "")
    return name.strip()


def sheet_cols(ws) -> dict[str, int]:
    return {clean(str(c.value or "")): i for i, c in enumerate(ws[1], start=1) if c.value}


def match_asset(stem: str, ids: list[str]) -> str:
    """Tên file bắt đầu bằng asset_id — chọn id KHỚP DÀI NHẤT để P-Q1-1 không nuốt P-Q1-18."""
    hits = [a for a in ids if stem.startswith(a)]
    return max(hits, key=len) if hits else ""


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--campaign-dir", required=True)
    ap.add_argument("--workbook", required=True)
    ap.add_argument("--utm-base")
    ap.add_argument("--utm-campaign")
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    cdir = Path(args.campaign_dir)
    cal = [r for r in load_calendar(args.workbook) if r.get("asset_id")]
    if not cal:
        print("03_Calendar rỗng — chưa có asset nào để gắn tài liệu.", file=sys.stderr)
        return 2
    ids = [r["asset_id"] for r in cal]
    by_id = {r["asset_id"]: r for r in cal}

    # ── quét file ────────────────────────────────────────────────────────────
    files, orphan = [], []
    for p in sorted(cdir.rglob("*")):
        if not p.is_file() or p.suffix.lower() not in KIND:
            continue
        if p.name.startswith("~$") or p.suffix.lower() == ".xlsx":
            continue
        rel = p.relative_to(cdir).as_posix()
        if rel in ("calendar.csv", "brief.yml"):
            continue
        # Ưu tiên THƯ MỤC theo mã content: assets/<asset_id>/... — tên file muốn đặt sao
        # cũng được. Chỉ khi file nằm ngoài cấu trúc đó mới đoán theo tiền tố tên file.
        parts = p.relative_to(cdir).parts
        aid = ""
        for seg in parts[:-1]:
            if seg in by_id:
                aid = seg
                break
        if not aid:
            aid = match_asset(p.stem, ids)
        if not aid:
            orphan.append(rel)
            continue
        role = next((v for k, v in ROLE_HINT.items() if k in p.stem.lower()), "other")
        data = p.read_bytes()
        files.append({
            "file_id": f"F-{hashlib.sha256(rel.encode()).hexdigest()[:8]}",
            "asset_id": aid, "file_type": KIND[p.suffix.lower()], "role": role,
            "rel_path": rel, "bytes": len(data), "dimensions": "", "duration_sec": "",
            "checksum": hashlib.sha256(data).hexdigest()[:16],
            # KHÔNG tự khai "own" — quyền là thứ người phải xác nhận.
            "rights_status": "unknown", "rights_source": "", "rights_expiry": "",
            "ai_generated": "", "created_at": datetime.fromtimestamp(
                p.stat().st_mtime).strftime("%Y-%m-%d %H:%M:%S"), "created_by": "",
        })

    # ── sinh link UTM ────────────────────────────────────────────────────────
    links = []
    if args.utm_base:
        camp = args.utm_campaign or (cal[0].get("campaign_id") if cal else "campaign")
        for r in cal:
            plat = (r.get("platform") or "").strip()
            if not plat:
                continue
            aid = r["asset_id"]
            src = plat
            med = UTM_MEDIUM.get(plat, "referral")
            url = (f"{args.utm_base}?utm_source={src}&utm_medium={med}"
                   f"&utm_campaign={camp}&utm_content={aid}")
            links.append({
                "link_id": f"L-{aid}", "asset_id": aid,
                "campaign_id": r.get("campaign_id", ""), "target_url": url,
                "utm_source": src, "utm_medium": med, "utm_campaign": camp,
                "utm_content": aid, "short_url": "", "landing_page": args.utm_base,
                "cta_label": r.get("cta_label") or r.get("cta_type", ""),
                "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            })

    print(f"── quét {cdir.name}")
    print(f"   {len(files)} file khớp asset · {len(links)} link UTM")
    unknown = sum(1 for f in files if f["rights_status"] == "unknown")
    if unknown:
        print(f"   ⚠ {unknown} file có rights_status = 'unknown'. Người PHẢI xác nhận quyền dùng")
        print(f"     trước khi publish — agent không được tự khai là 'own'.")
    if orphan:
        print(f"   ⚠ {len(orphan)} file không khớp asset_id nào (đặt tên sai?): "
              f"{', '.join(orphan[:4])}")
    for a in [i for i in ids if not any(f['asset_id'] == i for f in files)][:5]:
        print(f"   ℹ {a} chưa có file nào")
    if args.dry_run:
        print("   (dry-run — chưa ghi)")
        return 0

    wb = load_workbook(args.workbook)
    for sheet, rows in (("11_Assets", files), ("12_Links", links)):
        if sheet not in wb.sheetnames or not rows:
            continue
        ws = wb[sheet]
        cols = sheet_cols(ws)
        # giữ lại rights_* người đã điền, khoá theo rel_path / link_id
        keep = {}
        keyc = "rel_path" if sheet == "11_Assets" else "link_id"
        if keyc in cols:
            for r in range(2, ws.max_row + 1):
                k = ws.cell(r, cols[keyc]).value
                if k:
                    keep[str(k)] = {n: ws.cell(r, i).value for n, i in cols.items()}
        ws.delete_rows(2, ws.max_row)
        for j, row in enumerate(rows, start=2):
            old = keep.get(row[keyc], {})
            for name, i in cols.items():
                v = row.get(name, "")
                if name.startswith("rights_") and old.get(name) not in (None, "", "unknown"):
                    v = old[name]          # ← không đè xác nhận của người
                cell = ws.cell(j, i, v)
                if name == "rel_path" and v:          # bấm mở file ngay từ Excel
                    cell.hyperlink = str((cdir / str(v)).resolve())
                    cell.style = "Hyperlink"
                elif name == "target_url" and v:
                    cell.hyperlink = str(v)
                    cell.style = "Hyperlink"
        print(f"   ✅ {sheet}: {len(rows)} dòng")

    # 03_Calendar: cột asset_folder trỏ thẳng tới thư mục tài liệu của asset
    if "03_Calendar" in wb.sheetnames:
        ws = wb["03_Calendar"]
        cols = sheet_cols(ws)
        if "asset_folder" in cols and "asset_id" in cols:
            n = 0
            for r in range(2, ws.max_row + 1):
                aid = ws.cell(r, cols["asset_id"]).value
                if not aid:
                    continue
                folder = cdir / "assets" / str(aid)
                rel = f"assets/{aid}/"
                cell = ws.cell(r, cols["asset_folder"], rel)
                if folder.exists():
                    cell.hyperlink = str(folder.resolve())
                    cell.style = "Hyperlink"
                    n += 1
            print(f"   ✅ 03_Calendar.asset_folder: {n} asset đã có thư mục tài liệu")
    wb.save(args.workbook)
    return 0


if __name__ == "__main__":
    sys.exit(main())
