# -*- coding: utf-8 -*-
"""
CLI thao tác file Excel chiến dịch — mô hình 5 sheet KPIM.
Cổng vào/ra duy nhất giữa agent và workbook. Luật cổng ở schema/workbook_spec.yml.

    campaign_excel.py status        <wb.xlsx>
    campaign_excel.py get-campaign  <wb.xlsx>
    campaign_excel.py set-campaign  <wb.xlsx> --field F --value V
    campaign_excel.py list          <wb.xlsx> --stage draft [--json]
    campaign_excel.py upsert        <wb.xlsx> --json <post_row.json>
    campaign_excel.py set           <wb.xlsx> --post-id ID --field F --value V
    campaign_excel.py approve       <wb.xlsx> --post-id ID --gate approve_topic [--by "Duc"]
    campaign_excel.py result        <wb.xlsx> --post-id ID --json <result.json>
    campaign_excel.py add-asset     <wb.xlsx> --post-id ID --type T --rel-path R [--abs-path A]
    campaign_excel.py upsert-engagement <wb.xlsx> --json <eng_row.json>

Gating (list): draft=approve_topic&proposed · media=approve_content&drafted ·
preview=media_ready · publish=approve_final&{preview_ready,media_ready}.
Không tự tick approve thay user.

Exit: 0 ok · 1 vi phạm luật/không tìm thấy · 2 không chạy được.
"""
from __future__ import annotations

import argparse
import json
import os
import sys
from datetime import datetime
from pathlib import Path

try:
    import yaml
    from openpyxl import load_workbook
except ImportError as e:
    print(f"Thiếu thư viện: {e}. Cần: pip install pyyaml openpyxl", file=sys.stderr)
    sys.exit(2)

SPEC = Path(__file__).resolve().parents[2] / "schema" / "workbook_spec.yml"
TRUTHY = {"x", "true", "1", "✓", "✔", "yes", "có"}
GATES = ("approve_topic", "approve_content", "approve_final")


def spec() -> dict:
    return yaml.safe_load(SPEC.read_text(encoding="utf-8"))


def is_true(v) -> bool:
    return bool(v) and str(v).strip().lower() in TRUTHY


def now() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


class Table:
    def __init__(self, ws):
        self.ws = ws
        self.col = {str(c.value): i for i, c in enumerate(ws[1], 1) if c.value}

    def rows(self):
        for r in range(2, self.ws.max_row + 1):
            if self.ws.cell(r, 1).value in (None, ""):
                continue
            yield r, {n: self.ws.cell(r, i).value for n, i in self.col.items()}

    def find(self, key, val):
        for r, row in self.rows():
            if str(row.get(key) or "").strip() == str(val):
                return r
        return None

    def set(self, r, col, v):
        if col not in self.col:
            raise KeyError(f"sheet không có cột '{col}'")
        self.ws.cell(r, self.col[col], v)

    def append(self, values: dict) -> int:
        r = self.ws.max_row + 1
        while r > 2 and self.ws.cell(r - 1, 1).value in (None, ""):
            r -= 1
        for k, v in values.items():
            if k in self.col:
                self.ws.cell(r, self.col[k], v)
        return r


def open_wb(path):
    p = Path(path)
    if not p.exists():
        print(f"Không thấy workbook: {p}", file=sys.stderr)
        sys.exit(2)
    try:
        return load_workbook(p), p
    except PermissionError:
        print("Workbook đang mở trong Excel — đóng file rồi chạy lại.", file=sys.stderr)
        sys.exit(2)


def eligible(rows, stage, stages):
    st = stages[stage]
    out = []
    for r in rows:
        if st["from_status"] and str(r.get("status") or "").strip() not in st["from_status"]:
            continue
        gate = st.get("gate")
        if gate and not is_true(r.get(gate)):
            continue
        out.append(r)
    return out


# ── lệnh ────────────────────────────────────────────────────────────────────
def cmd_status(args):
    wb, p = open_wb(args.workbook)
    post = [r for _, r in Table(wb["Post"]).rows()]
    stages = spec()["stages"]
    print(f"── {p.name} · {len(post)} bài")
    from collections import Counter
    st = Counter(str(r.get("status") or "?") for r in post)
    print("   status: " + " · ".join(f"{k}={v}" for k, v in sorted(st.items())))
    for g in GATES:
        print(f"   {g}: {sum(1 for r in post if is_true(r.get(g)))}/{len(post)} đã tick")
    print("   sẵn sàng theo stage:")
    for name in stages:
        if name == "topics":
            continue
        print(f"      {name:<8} {len(eligible(post, name, stages))}")
    return 0


def cmd_get_campaign(args):
    wb, _ = open_wb(args.workbook)
    ws = wb["Campaign"]
    out = {ws.cell(r, 1).value: ws.cell(r, 2).value
           for r in range(2, ws.max_row + 1) if ws.cell(r, 1).value}
    print(json.dumps(out, ensure_ascii=False, indent=2, default=str))
    return 0


def cmd_set_campaign(args):
    wb, p = open_wb(args.workbook)
    ws = wb["Campaign"]
    for r in range(2, ws.max_row + 1):
        if str(ws.cell(r, 1).value) == args.field:
            ws.cell(r, 2, args.value)
            wb.save(p)
            print(f"✅ Campaign.{args.field} = {args.value}")
            return 0
    print(f"Không có field '{args.field}' trong Sheet Campaign", file=sys.stderr)
    return 1


def cmd_list(args):
    wb, _ = open_wb(args.workbook)
    rows = [r for _, r in Table(wb["Post"]).rows()]
    stages = spec()["stages"]
    if args.stage not in stages:
        print(f"stage không hợp lệ: {args.stage}. Có: {', '.join(stages)}", file=sys.stderr)
        return 2
    ok = eligible(rows, args.stage, stages)
    if args.json:
        print(json.dumps(ok, ensure_ascii=False, indent=2, default=str))
        return 0
    st = stages[args.stage]
    print(f"── stage '{args.stage}' · status ∈ {st['from_status'] or 'mọi'} · cổng: {st.get('gate') or '—'}")
    if not ok:
        blocked = [r for r in rows if str(r.get("status") or "") in st["from_status"]]
        if blocked and st.get("gate"):
            print(f"   0 bài. {len(blocked)} bài đúng status nhưng CHƯA tick '{st['gate']}':")
            for r in blocked[:8]:
                print(f"      {r.get('post_id')} — {str(r.get('topic_title') or '')[:52]}")
            print("   → Người phải tick cột đó (hoặc dùng lệnh approve).")
        else:
            print("   0 bài đủ điều kiện.")
        return 0
    for r in ok:
        print(f"   {str(r.get('post_id')):<12} {str(r.get('topic_title') or '')[:56]}")
    print(f"   → {len(ok)} bài. Xử lý xong đặt status = '{st['to_status']}'.")
    return 0


def cmd_upsert(args):
    wb, p = open_wb(args.workbook)
    row = json.loads(Path(args.json).read_text(encoding="utf-8")) if Path(args.json).exists() \
        else json.loads(args.json)
    t = Table(wb["Post"])
    pid = row.get("post_id")
    r = t.find("post_id", pid)
    if r:
        for k, v in row.items():
            if k in t.col:
                t.set(r, k, v)
        action = "cập nhật"
    else:
        t.append(row)
        action = "thêm"
    wb.save(p)
    print(f"✅ {action} Post: {pid}")
    return 0


def cmd_set(args):
    if args.field in GATES:
        print(f"✖ '{args.field}' là cổng duyệt của NGƯỜI. Dùng lệnh approve, hoặc tick trong Excel.",
              file=sys.stderr)
        return 1
    wb, p = open_wb(args.workbook)
    t = Table(wb["Post"])
    r = t.find("post_id", args.post_id)
    if r is None:
        print(f"Không thấy post {args.post_id}", file=sys.stderr)
        return 1
    t.set(r, args.field, args.value)
    wb.save(p)
    print(f"✅ {args.post_id}.{args.field} = {args.value}")
    return 0


def cmd_approve(args):
    if args.gate not in GATES:
        print(f"gate phải là một trong {GATES}", file=sys.stderr)
        return 1
    wb, p = open_wb(args.workbook)
    t = Table(wb["Post"])
    r = t.find("post_id", args.post_id)
    if r is None:
        print(f"Không thấy post {args.post_id}", file=sys.stderr)
        return 1
    t.set(r, args.gate, "x")
    wb.save(p)
    print(f"✅ {args.post_id}: tick {args.gate}" + (f" bởi {args.by}" if args.by else ""))
    return 0


def cmd_result(args):
    wb, p = open_wb(args.workbook)
    row = json.loads(Path(args.json).read_text(encoding="utf-8")) if Path(args.json).exists() \
        else json.loads(args.json)
    row["post_id"] = args.post_id
    t = Table(wb["Result"])
    r = t.find("post_id", args.post_id)
    if r:
        for k, v in row.items():
            if k in t.col:
                t.set(r, k, v)
    else:
        t.append(row)
    wb.save(p)
    print(f"✅ Result: {args.post_id}")
    return 0


def cmd_add_asset(args):
    wb, p = open_wb(args.workbook)
    t = Table(wb["Assets"])
    t.append({"post_id": args.post_id, "asset_type": args.type,
              "rel_path": args.rel_path, "abs_path": args.abs_path or "",
              "size": "", "created": now()})
    wb.save(p)
    print(f"✅ Assets: {args.post_id} · {args.type} · {args.rel_path}")
    return 0


def cmd_upsert_eng(args):
    wb, p = open_wb(args.workbook)
    row = json.loads(Path(args.json).read_text(encoding="utf-8")) if Path(args.json).exists() \
        else json.loads(args.json)
    row.setdefault("fetched_at", now())
    t = Table(wb["Engagement"])
    r = t.find("post_id", row.get("post_id"))
    if r:
        for k, v in row.items():
            if k in t.col:
                t.set(r, k, v)
    else:
        t.append(row)
    wb.save(p)
    print(f"✅ Engagement: {row.get('post_id')}")
    return 0


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    sub = ap.add_subparsers(dest="cmd", required=True)

    def add(name, fn, *fields):
        p = sub.add_parser(name)
        p.add_argument("workbook")
        for f in fields:
            p.add_argument(f, required=True)
        p.set_defaults(fn=fn)
        return p

    add("status", cmd_status)
    add("get-campaign", cmd_get_campaign)
    add("set-campaign", cmd_set_campaign, "--field", "--value")
    p = add("list", cmd_list, "--stage"); p.add_argument("--json", action="store_true")
    add("upsert", cmd_upsert, "--json")
    add("set", cmd_set, "--post-id", "--field", "--value")
    p = add("approve", cmd_approve, "--post-id", "--gate"); p.add_argument("--by")
    add("result", cmd_result, "--post-id", "--json")
    p = add("add-asset", cmd_add_asset, "--post-id", "--type", "--rel-path"); p.add_argument("--abs-path")
    add("upsert-engagement", cmd_upsert_eng, "--json")

    args = ap.parse_args()
    return args.fn(args)


if __name__ == "__main__":
    sys.exit(main())
