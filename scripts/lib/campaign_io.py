# -*- coding: utf-8 -*-
"""
Đọc dữ liệu chiến dịch — mô hình 5 sheet KPIM. Workbook là nguồn sự thật.
Mọi script dùng chung bộ đọc này để không lệch nhau.
"""
from __future__ import annotations

from pathlib import Path

MARKS = ("🤖", "👤", "⚙️", "⚙")


def _clean(name) -> str:
    s = str(name or "")
    for m in MARKS:
        s = s.replace(m, "")
    return s.strip()


def read_sheet(workbook: str | Path, sheet: str) -> list[dict]:
    """Đọc một sheet bảng thành list[dict]. Bỏ dòng trống ở cột đầu."""
    from openpyxl import load_workbook
    wb = load_workbook(workbook, read_only=True, data_only=True)
    try:
        if sheet not in wb.sheetnames:
            return []
        it = wb[sheet].iter_rows(values_only=True)
        try:
            hdr = [_clean(h) for h in next(it)]
        except StopIteration:
            return []
        out = []
        for row in it:
            if not row or row[0] in (None, ""):
                continue
            out.append({h: ("" if v is None else v) for h, v in zip(hdr, row) if h})
        return out
    finally:
        wb.close()


def read_campaign(workbook: str | Path) -> dict:
    """Sheet Campaign là FORM dọc (field | value | chú thích) → dict field→value."""
    from openpyxl import load_workbook
    wb = load_workbook(workbook, read_only=True, data_only=True)
    try:
        if "Campaign" not in wb.sheetnames:
            return {}
        out = {}
        for row in wb["Campaign"].iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                out[_clean(row[0])] = "" if len(row) < 2 or row[1] is None else row[1]
        return out
    finally:
        wb.close()


def find_workbook(campaign_dir: str | Path) -> Path | None:
    """File .xlsx duy nhất của chiến dịch. Nhiều hơn một = sai quy ước → báo lỗi."""
    d = Path(campaign_dir)
    hits = [f for f in d.glob("*.xlsx") if not f.name.startswith("~$")]
    if len(hits) > 1:
        raise RuntimeError(f"{d.name} có {len(hits)} file .xlsx — một chiến dịch chỉ được MỘT.")
    return hits[0] if hits else None
